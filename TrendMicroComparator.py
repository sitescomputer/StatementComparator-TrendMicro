#!/usr/bin/env python3
"""Compare two monthly Trend Micro MSP customer-summary workbooks.

Run without arguments to open the Windows-friendly Tkinter application. The
comparison engine is intentionally kept separate from the GUI so it can also be
tested, scripted, and packaged with PyInstaller.
"""

from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass, field
from datetime import datetime
import os
from pathlib import Path
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from typing import Iterable, Sequence

from openpyxl import load_workbook


APP_NAME = "Trend Micro Billing Comparator"

OUTPUT_COLUMNS = (
    "Customer ID",
    "Customer",
    "City",
    "State",
    "Old Service Plan",
    "New Service Plan",
    "Old Provisioned",
    "New Provisioned",
    "Provisioned Change",
    "Old Used Seats",
    "New Used Seats",
    "Used Seats Change",
    "Status",
)

REQUIRED_FIELDS = ("customer", "service_plan", "provisioned", "used")
HEADER_ALIASES = {
    "customer_id": {"customer id", "customerid"},
    "customer": {"customer", "customer name"},
    "city": {"city"},
    "state": {"state", "province", "state province"},
    "service_plan": {"service plan", "plan", "serviceplan"},
    "provisioned": {
        "provisioned",
        "provisioned seats",
        "provisioned units",
        "provisioned quantity",
    },
    "used": {"used", "used seats", "used units", "seats used"},
}


class TrendMicroFormatError(ValueError):
    """Raised when an input workbook is not a recognized customer summary."""


@dataclass(frozen=True)
class CustomerRecord:
    """One customer/service-plan row from a Trend Micro workbook."""

    customer: str
    service_plan: str
    provisioned: int | float
    used: int | float
    customer_id: str = ""
    city: str = ""
    state: str = ""
    source_row: int = 0

    @property
    def customer_key(self) -> str:
        return _normalize_customer_name(self.customer)

    @property
    def customer_id_key(self) -> str:
        return _normalize_comparison_text(self.customer_id)

    @property
    def plan_key(self) -> str:
        return _normalize_comparison_text(self.service_plan)


@dataclass(frozen=True)
class WorkbookData:
    """Parsed customer records and the worksheet they came from."""

    path: Path
    sheet_name: str
    records: tuple[CustomerRecord, ...]


@dataclass
class ComparisonReport:
    """Comparison rows plus summary information for the GUI and CLI."""

    old_path: Path
    new_path: Path
    old_sheet: str
    new_sheet: str
    old_record_count: int
    new_record_count: int
    rows: list[dict[str, object]] = field(default_factory=list)
    output_path: Path | None = None

    @property
    def compared_count(self) -> int:
        return len(self.rows)

    @property
    def changed_count(self) -> int:
        return sum(row["Status"] != "-" for row in self.rows)

    @property
    def category_counts(self) -> Counter[str]:
        counts: Counter[str] = Counter()
        for row in self.rows:
            status = str(row["Status"])
            if status == "-":
                continue
            counts.update(part.strip() for part in status.split(";") if part.strip())
        return counts


def _clean_display_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ")).strip()


def _normalize_comparison_text(value: object) -> str:
    text = unicodedata.normalize("NFKC", _clean_display_text(value))
    text = text.translate(str.maketrans({"’": "'", "‘": "'", "™": "", "®": ""}))
    return re.sub(r"\s+", " ", text).strip().casefold()


def _normalize_customer_name(value: object) -> str:
    text = _normalize_comparison_text(value)
    # Treat harmless punctuation spacing differences as the same customer name.
    return re.sub(r"\s*([,&])\s*", r"\1", text)


def _normalize_header(value: object) -> str:
    text = _normalize_comparison_text(value)
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _field_for_header(value: object) -> str | None:
    normalized = _normalize_header(value)
    for field_name, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return field_name
    return None


def _coerce_seat_count(value: object, field_name: str, row_number: int) -> int | float:
    if value is None or _clean_display_text(value) == "":
        return 0
    if isinstance(value, bool):
        raise TrendMicroFormatError(
            f"Row {row_number}: {field_name} must be a seat count, not {value!r}."
        )
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        cleaned = _clean_display_text(value).replace(",", "")
        try:
            number = float(cleaned)
        except ValueError as exc:
            raise TrendMicroFormatError(
                f"Row {row_number}: {field_name} contains a non-numeric value: {value!r}."
            ) from exc
    if number.is_integer():
        return int(number)
    return number


def _find_header_row(worksheet, max_rows: int = 50) -> tuple[int, dict[str, int]] | None:
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1
    ):
        columns: dict[str, int] = {}
        for column_index, value in enumerate(row):
            field_name = _field_for_header(value)
            if field_name is not None and field_name not in columns:
                columns[field_name] = column_index
        if all(field in columns for field in REQUIRED_FIELDS):
            return row_number, columns
    return None


def read_trend_micro_workbook(path: str | Path) -> WorkbookData:
    """Read the first worksheet containing the required Trend Micro columns."""

    workbook_path = Path(path).expanduser().resolve()
    if workbook_path.suffix.casefold() != ".xlsx":
        raise TrendMicroFormatError(
            f"{workbook_path.name} is not an .xlsx workbook. Export or save it as .xlsx first."
        )
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        raise TrendMicroFormatError(
            f"Could not open {workbook_path.name} as an Excel workbook: {exc}"
        ) from exc

    try:
        selected = None
        for worksheet in workbook.worksheets:
            header = _find_header_row(worksheet)
            if header is not None:
                selected = (worksheet, *header)
                break

        if selected is None:
            required = "Customer, Service Plan, Provisioned, and Used"
            raise TrendMicroFormatError(
                f"{workbook_path.name} does not contain a worksheet with the required "
                f"columns: {required}. The header may appear within the first 50 rows."
            )

        worksheet, header_row, columns = selected
        records: list[CustomerRecord] = []
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            customer = _clean_display_text(row[columns["customer"]])
            if not customer:
                # This excludes the total rows at the bottom of complete exports.
                continue

            def optional_text(field_name: str) -> str:
                column = columns.get(field_name)
                if column is None or column >= len(row):
                    return ""
                return _clean_display_text(row[column])

            records.append(
                CustomerRecord(
                    customer=customer,
                    customer_id=optional_text("customer_id"),
                    city=optional_text("city"),
                    state=optional_text("state"),
                    service_plan=optional_text("service_plan"),
                    provisioned=_coerce_seat_count(
                        row[columns["provisioned"]], "Provisioned", row_number
                    ),
                    used=_coerce_seat_count(row[columns["used"]], "Used", row_number),
                    source_row=row_number,
                )
            )

        if not records:
            raise TrendMicroFormatError(
                f"{workbook_path.name} has the expected headers but no customer rows."
            )
        return WorkbookData(
            path=workbook_path,
            sheet_name=worksheet.title,
            records=tuple(records),
        )
    finally:
        workbook.close()


def _pair_bucket(
    old_indices: Iterable[int],
    new_indices: Iterable[int],
    old_records: Sequence[CustomerRecord],
    new_records: Sequence[CustomerRecord],
) -> list[tuple[int, int]]:
    """Pair duplicate customer rows, preferring an unchanged plan first."""

    old_remaining = set(old_indices)
    new_remaining = set(new_indices)
    pairs: list[tuple[int, int]] = []

    old_by_plan: dict[str, list[int]] = defaultdict(list)
    new_by_plan: dict[str, list[int]] = defaultdict(list)
    for index in old_remaining:
        old_by_plan[old_records[index].plan_key].append(index)
    for index in new_remaining:
        new_by_plan[new_records[index].plan_key].append(index)

    for plan_key in sorted(set(old_by_plan) & set(new_by_plan)):
        for old_index, new_index in zip(
            sorted(old_by_plan[plan_key]), sorted(new_by_plan[plan_key])
        ):
            pairs.append((old_index, new_index))
            old_remaining.remove(old_index)
            new_remaining.remove(new_index)

    # Remaining rows represent likely plan changes. Pair deterministically so
    # one changed plan is not reported as both a dropped and a new customer.
    old_sorted = sorted(
        old_remaining,
        key=lambda index: (
            old_records[index].plan_key,
            old_records[index].source_row,
        ),
    )
    new_sorted = sorted(
        new_remaining,
        key=lambda index: (
            new_records[index].plan_key,
            new_records[index].source_row,
        ),
    )
    pairs.extend(zip(old_sorted, new_sorted))
    return pairs


def _match_records(
    old_records: Sequence[CustomerRecord], new_records: Sequence[CustomerRecord]
) -> tuple[list[tuple[int, int]], set[int], set[int]]:
    old_unmatched = set(range(len(old_records)))
    new_unmatched = set(range(len(new_records)))
    pairs: list[tuple[int, int]] = []

    def pair_by(key_getter) -> None:
        old_buckets: dict[str, list[int]] = defaultdict(list)
        new_buckets: dict[str, list[int]] = defaultdict(list)
        for index in old_unmatched:
            key = key_getter(old_records[index])
            if key:
                old_buckets[key].append(index)
        for index in new_unmatched:
            key = key_getter(new_records[index])
            if key:
                new_buckets[key].append(index)

        for key in sorted(set(old_buckets) & set(new_buckets)):
            bucket_pairs = _pair_bucket(
                old_buckets[key], new_buckets[key], old_records, new_records
            )
            for old_index, new_index in bucket_pairs:
                pairs.append((old_index, new_index))
                old_unmatched.remove(old_index)
                new_unmatched.remove(new_index)

    # Stable Trend Micro customer IDs win when both months contain them. The
    # name pass handles carved files where the ID column was removed.
    pair_by(lambda record: record.customer_id_key)
    pair_by(lambda record: record.customer_key)

    return pairs, old_unmatched, new_unmatched


def _format_comparison_row(
    old: CustomerRecord | None,
    new: CustomerRecord | None,
    status: str,
) -> dict[str, object]:
    current = new or old
    assert current is not None

    if new is not None and old is None:
        provisioned_change: object = new.provisioned
        used_change: object = new.used
    elif old is not None and new is not None:
        provisioned_change = new.provisioned - old.provisioned
        used_change = new.used - old.used
    else:
        # Match AutoMailprotector.py: new-side and change fields are blank for
        # dropped entries while their previous values remain visible.
        provisioned_change = ""
        used_change = ""

    return {
        "Customer ID": (new.customer_id if new else "") or (old.customer_id if old else ""),
        "Customer": current.customer,
        "City": current.city,
        "State": current.state,
        "Old Service Plan": old.service_plan if old else "",
        "New Service Plan": new.service_plan if new else "",
        "Old Provisioned": old.provisioned if old else "",
        "New Provisioned": new.provisioned if new else "",
        "Provisioned Change": provisioned_change,
        "Old Used Seats": old.used if old else "",
        "New Used Seats": new.used if new else "",
        "Used Seats Change": used_change,
        "Status": status,
    }


def compare_records(
    old_records: Sequence[CustomerRecord], new_records: Sequence[CustomerRecord]
) -> list[dict[str, object]]:
    """Return a full outer comparison of two monthly customer lists."""

    pairs, old_unmatched, new_unmatched = _match_records(old_records, new_records)
    rows: list[dict[str, object]] = []

    for old_index, new_index in pairs:
        old = old_records[old_index]
        new = new_records[new_index]
        statuses: list[str] = []
        if old.customer_key != new.customer_key:
            statuses.append("CUSTOMER NAME CHANGE")
        if old.plan_key != new.plan_key:
            statuses.append("SERVICE PLAN CHANGE")
        if old.provisioned != new.provisioned:
            statuses.append("PROVISIONED CHANGE")
        if old.used != new.used:
            statuses.append("USED SEATS CHANGE")
        rows.append(_format_comparison_row(old, new, "; ".join(statuses) or "-"))

    old_customer_ids = {record.customer_id_key for record in old_records if record.customer_id_key}
    new_customer_ids = {record.customer_id_key for record in new_records if record.customer_id_key}
    old_customer_names = {record.customer_key for record in old_records}
    new_customer_names = {record.customer_key for record in new_records}

    for new_index in new_unmatched:
        new = new_records[new_index]
        existed_before = (
            (new.customer_id_key and new.customer_id_key in old_customer_ids)
            or new.customer_key in old_customer_names
        )
        status = "NEW SERVICE PLAN" if existed_before else "NEW CUSTOMER"
        rows.append(_format_comparison_row(None, new, status))

    for old_index in old_unmatched:
        old = old_records[old_index]
        still_exists = (
            (old.customer_id_key and old.customer_id_key in new_customer_ids)
            or old.customer_key in new_customer_names
        )
        status = "DROPPED SERVICE PLAN" if still_exists else "DROPPED CUSTOMER"
        rows.append(_format_comparison_row(old, None, status))

    status_order = {
        "NEW CUSTOMER": 0,
        "DROPPED CUSTOMER": 1,
        "NEW SERVICE PLAN": 2,
        "DROPPED SERVICE PLAN": 3,
        "SERVICE PLAN CHANGE": 4,
        "PROVISIONED CHANGE": 5,
        "USED SEATS CHANGE": 6,
        "CUSTOMER NAME CHANGE": 7,
        "-": 99,
    }

    def row_sort_key(row: dict[str, object]) -> tuple[object, ...]:
        status = str(row["Status"])
        first_status = status.split(";", 1)[0]
        return (
            _normalize_customer_name(row["Customer"]),
            status_order.get(first_status, 50),
            _normalize_comparison_text(row["New Service Plan"] or row["Old Service Plan"]),
        )

    rows.sort(key=row_sort_key)
    return rows


def write_comparison_csv(rows: Sequence[dict[str, object]], output_path: str | Path) -> Path:
    """Write an Excel-friendly UTF-8 CSV containing all comparison rows."""

    destination = Path(output_path).expanduser().resolve()
    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=OUTPUT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            # Prevent text supplied by a workbook from becoming an Excel formula
            # when the CSV is opened. Numeric values remain numeric.
            safe_row: dict[str, object] = {}
            for key, value in row.items():
                stripped = value.lstrip() if isinstance(value, str) else ""
                formula_like = stripped.startswith(("=", "+", "@")) or (
                    stripped.startswith("-") and stripped != "-"
                )
                safe_row[key] = f"'{value}" if formula_like else value
            writer.writerow(safe_row)
    return destination


def compare_workbooks(
    old_path: str | Path,
    new_path: str | Path,
    output_path: str | Path,
) -> ComparisonReport:
    """Parse, compare, and save two Trend Micro monthly workbooks."""

    old_data = read_trend_micro_workbook(old_path)
    new_data = read_trend_micro_workbook(new_path)
    rows = compare_records(old_data.records, new_data.records)
    destination = write_comparison_csv(rows, output_path)
    return ComparisonReport(
        old_path=old_data.path,
        new_path=new_data.path,
        old_sheet=old_data.sheet_name,
        new_sheet=new_data.sheet_name,
        old_record_count=len(old_data.records),
        new_record_count=len(new_data.records),
        rows=rows,
        output_path=destination,
    )


def default_output_path() -> Path:
    downloads = Path.home() / "Downloads"
    output_directory = downloads if downloads.is_dir() else Path.home()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return output_directory / f"TrendMicro_Billing_Comparison_{timestamp}.csv"


def _status_summary(report: ComparisonReport) -> str:
    labels = (
        "NEW CUSTOMER",
        "DROPPED CUSTOMER",
        "SERVICE PLAN CHANGE",
        "PROVISIONED CHANGE",
        "USED SEATS CHANGE",
        "CUSTOMER NAME CHANGE",
        "NEW SERVICE PLAN",
        "DROPPED SERVICE PLAN",
    )
    counts = report.category_counts
    details = [f"{label}: {counts[label]}" for label in labels if counts[label]]
    return "\n".join(details) if details else "No differences found."


def _run_gui() -> int:
    """Run the Qt desktop interface."""

    from PySide6.QtCore import QObject, QThread, QTimer, QUrl, Signal, Slot
    from PySide6.QtGui import QDesktopServices, QFont
    from PySide6.QtWidgets import (
        QApplication,
        QFileDialog,
        QGroupBox,
        QHBoxLayout,
        QLabel,
        QMessageBox,
        QProgressBar,
        QPushButton,
        QVBoxLayout,
        QWidget,
    )

    class ComparisonWorker(QObject):
        finished = Signal(object)
        failed = Signal(str)

        def __init__(self, old_path: Path, new_path: Path) -> None:
            super().__init__()
            self.old_path = old_path
            self.new_path = new_path

        @Slot()
        def run(self) -> None:
            try:
                report = compare_workbooks(
                    self.old_path,
                    self.new_path,
                    default_output_path(),
                )
            except Exception as exc:
                self.failed.emit(str(exc))
                return
            self.finished.emit(report)

    class TrendMicroComparisonWindow(QWidget):
        def __init__(self) -> None:
            super().__init__()
            self.old_month_file: Path | None = None
            self.new_month_file: Path | None = None
            self.output_path: Path | None = None
            self.worker_thread: QThread | None = None
            self.worker: ComparisonWorker | None = None

            self.setWindowTitle(APP_NAME)
            self.resize(720, 570)
            self.setMinimumSize(680, 540)
            self._setup_gui()

        def _setup_gui(self) -> None:
            main = QVBoxLayout(self)
            main.setContentsMargins(24, 24, 24, 24)
            main.setSpacing(12)

            title = QLabel(APP_NAME)
            title.setFont(QFont("Segoe UI", 18, QFont.Weight.Bold))
            main.addWidget(title)

            instructions = QLabel(
                "Select the older month first and the newer month second. Complete Trend Micro "
                "exports and carved-down files are both supported."
            )
            instructions.setWordWrap(True)
            main.addWidget(instructions)

            old_group = QGroupBox("Older Month")
            old_layout = QHBoxLayout(old_group)
            self.old_label = QLabel("No file selected")
            self.old_label.setFrameStyle(QLabel.Shape.Panel | QLabel.Shadow.Sunken)
            self.old_label.setMargin(5)
            old_button = QPushButton("Browse...")
            old_button.clicked.connect(self._browse_old)
            old_layout.addWidget(self.old_label, 1)
            old_layout.addWidget(old_button)
            main.addWidget(old_group)

            new_group = QGroupBox("Newer Month")
            new_layout = QHBoxLayout(new_group)
            self.new_label = QLabel("No file selected")
            self.new_label.setFrameStyle(QLabel.Shape.Panel | QLabel.Shadow.Sunken)
            self.new_label.setMargin(5)
            new_button = QPushButton("Browse...")
            new_button.clicked.connect(self._browse_new)
            new_layout.addWidget(self.new_label, 1)
            new_layout.addWidget(new_button)
            main.addWidget(new_group)

            self.compare_button = QPushButton("Compare Months")
            self.compare_button.setEnabled(False)
            self.compare_button.clicked.connect(self._start_comparison)
            main.addWidget(self.compare_button, 0)

            self.progress = QProgressBar()
            self.progress.setRange(0, 0)
            self.progress.hide()
            main.addWidget(self.progress)

            self.status_label = QLabel("")
            self.status_label.setWordWrap(True)
            main.addWidget(self.status_label)

            self.output_group = QGroupBox("Comparison Results")
            output_layout = QVBoxLayout(self.output_group)
            self.output_info = QLabel("")
            self.output_info.setWordWrap(True)
            self.output_info.setTextInteractionFlags(
                self.output_info.textInteractionFlags()
            )
            output_layout.addWidget(self.output_info)
            self.open_output_button = QPushButton("Open Output File")
            self.open_output_button.clicked.connect(self._open_output_file)
            self.open_output_button.setEnabled(False)
            output_layout.addWidget(self.open_output_button)
            self.output_group.hide()
            main.addWidget(self.output_group, 1)

        def _browse_old(self) -> None:
            filename, _ = QFileDialog.getOpenFileName(
                self,
                "Select the Older Month Trend Micro Workbook",
                "",
                "Excel workbooks (*.xlsx);;All files (*.*)",
            )
            if filename:
                self.old_month_file = Path(filename)
                self.old_label.setText(self.old_month_file.name)
                self._check_ready()

        def _browse_new(self) -> None:
            filename, _ = QFileDialog.getOpenFileName(
                self,
                "Select the Newer Month Trend Micro Workbook",
                "",
                "Excel workbooks (*.xlsx);;All files (*.*)",
            )
            if filename:
                self.new_month_file = Path(filename)
                self.new_label.setText(self.new_month_file.name)
                self._check_ready()

        def _check_ready(self) -> None:
            self.compare_button.setEnabled(
                self.old_month_file is not None and self.new_month_file is not None
            )

        def _set_status(self, message: str, color: str = "") -> None:
            self.status_label.setText(message)
            self.status_label.setStyleSheet(f"color: {color};" if color else "")

        def _start_comparison(self) -> None:
            assert self.old_month_file is not None and self.new_month_file is not None
            if self.old_month_file.resolve() == self.new_month_file.resolve():
                QMessageBox.warning(self, "Select Two Months", "Please select two different workbooks.")
                return

            self.compare_button.setEnabled(False)
            self.output_path = None
            self.open_output_button.setEnabled(False)
            self.output_group.hide()
            self.progress.show()
            self._set_status("Reading and comparing the two Trend Micro workbooks...")

            self.worker_thread = QThread(self)
            self.worker = ComparisonWorker(self.old_month_file, self.new_month_file)
            self.worker.moveToThread(self.worker_thread)
            self.worker_thread.started.connect(self.worker.run)
            self.worker.finished.connect(self._show_results)
            self.worker.failed.connect(self._show_error)
            self.worker.finished.connect(self.worker_thread.quit)
            self.worker.failed.connect(self.worker_thread.quit)
            self.worker_thread.finished.connect(self.worker.deleteLater)
            self.worker_thread.finished.connect(self._thread_finished)
            self.worker_thread.start()

        @Slot()
        def _thread_finished(self) -> None:
            assert self.worker_thread is not None
            self.worker_thread.deleteLater()
            self.worker_thread = None
            self.worker = None

        def _finish_progress(self) -> None:
            self.progress.hide()
            self.compare_button.setEnabled(True)

        @Slot(object)
        def _show_results(self, report: ComparisonReport) -> None:
            self._finish_progress()
            self._set_status("Comparison completed successfully.", "green")
            self.output_path = report.output_path
            self.open_output_button.setEnabled(self.output_path is not None)
            self.output_group.show()
            self.output_info.setText(
                f"Older month: {report.old_record_count} customer/service rows\n"
                f"Newer month: {report.new_record_count} customer/service rows\n"
                f"Rows with differences: {report.changed_count}\n\n"
                f"{_status_summary(report)}\n\n"
                f"Output saved to:\n{report.output_path}"
            )

        @Slot()
        def _open_output_file(self) -> None:
            if self.output_path is None:
                return
            opened = QDesktopServices.openUrl(
                QUrl.fromLocalFile(str(self.output_path.resolve()))
            )
            if not opened:
                QMessageBox.warning(
                    self,
                    "Could Not Open File",
                    f"Windows could not open the file automatically.\n\n{self.output_path}",
                )

        @Slot(str)
        def _show_error(self, error_message: str) -> None:
            self._finish_progress()
            self._set_status("The comparison could not be completed.", "red")
            QMessageBox.critical(self, "Comparison Error", error_message)

    application = QApplication([sys.argv[0]])
    application.setApplicationName(APP_NAME)
    window = TrendMicroComparisonWindow()
    window.show()
    screen = application.primaryScreen()
    if screen is not None:
        frame = window.frameGeometry()
        frame.moveCenter(screen.availableGeometry().center())
        window.move(frame.topLeft())
    if os.environ.get("TM_COMPARATOR_GUI_SMOKE_TEST") == "1":
        QTimer.singleShot(250, application.quit)
    return application.exec()


def _build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--old", help="Path to the older month's .xlsx workbook")
    parser.add_argument("--new", help="Path to the newer month's .xlsx workbook")
    parser.add_argument("--output", help="Output CSV path (defaults to Downloads)")
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = _build_argument_parser()
    args = parser.parse_args(argv)
    if args.old or args.new or args.output:
        if not args.old or not args.new:
            parser.error("--old and --new must be supplied together")
        output = Path(args.output) if args.output else default_output_path()
        try:
            report = compare_workbooks(args.old, args.new, output)
        except Exception as exc:
            print(f"Error: {exc}", file=sys.stderr)
            return 1
        print(
            f"Compared {report.old_record_count} older rows with "
            f"{report.new_record_count} newer rows."
        )
        print(f"Rows with differences: {report.changed_count}")
        print(_status_summary(report))
        print(f"Saved: {report.output_path}")
        return 0
    return _run_gui()


if __name__ == "__main__":
    raise SystemExit(main())
